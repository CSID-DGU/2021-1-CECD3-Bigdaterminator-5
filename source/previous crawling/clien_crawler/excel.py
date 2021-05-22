import openpyxl

def make_excel(file_name):
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    header = ['No.', '종류', '작성자', '조회수', '작성 날짜', '제목', '링크', '내용 / 댓글']
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 7
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 60
    ws.column_dimensions['H'].width = 100
    ws.append(header)
    wb.save(file_name)

def append_excel(file_name, item, start_idx=1):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.worksheets[0]

    idx = start_idx * 1
    for it in item:
        row = [idx, '게시글', it['author'], it['view'], it['date'], it['title'], it['link'], it['contents']]
        ws.append(row)
        idx += 1
        # 댓글 추가
        for comment in it['comments']:
            row2 = ['', '댓글', comment['author'], '', comment['date'], '', '', comment['comment']]
            ws.append(row2)
    
    wb.save(file_name)

    return idx + 1

def make_excel2(file_name):
    wb2 = openpyxl.Workbook()
    ws2 = wb2.worksheets[0]
    header = ['No.', '내용']
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 50
    ws2.append(header)
    wb2.save(file_name)

def append_excel2(file_name, item, start_idx2=1):
    wb2 = openpyxl.load_workbook(file_name)
    ws2 = wb2.worksheets[0]

    idxw = start_idx2 * 1
    for i in item:
        row = [idxw, i['morphs1'], i['morphs2'], i['counts'], i['morphs4']]
        ws2.append(row)
        idxw += 1
        # 댓글 추가
        for comment in i['comments']:
            row2 = ['', comment['morphs3'], '', comment['counts']]
            ws2.append(row2)
    
    wb2.save(file_name)

    return idxw + 1